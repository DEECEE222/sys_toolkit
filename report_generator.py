#!/usr/bin/env python3
"""
report_generator.py - Generación de informes ejecutivos en Excel.
Usa Pandas + OpenPyXL para crear reportes profesionales con formato.
"""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter

from inventory_manager import (
    load_inventory,
    filter_windows_servers,
    filter_low_ram,
    filter_vulnerable_servers,
    group_by_department,
)


# ─── Paleta de colores corporativa ───────────────────────────────────────────
COLOR_HEADER_BG = "1F3864"   # Azul oscuro
COLOR_HEADER_FG = "FFFFFF"   # Blanco
COLOR_ALERT_BG  = "FF4444"   # Rojo alerta
COLOR_WARN_BG   = "FFA500"   # Naranja advertencia
COLOR_OK_BG     = "4CAF50"   # Verde OK
COLOR_ROW_ALT   = "EBF0FA"   # Azul claro alternado
COLOR_TITLE_BG  = "2E75B6"   # Azul título


def _apply_header_style(ws, row: int, num_cols: int) -> None:
    """Aplica estilo de cabecera a una fila."""
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(color=COLOR_HEADER_FG, bold=True, size=10)
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def _apply_data_rows_style(ws, start_row: int, end_row: int, num_cols: int) -> None:
    """Aplica estilo alternado a las filas de datos."""
    alt_fill = PatternFill("solid", fgColor=COLOR_ROW_ALT)
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in range(start_row, end_row + 1):
        use_alt = (row - start_row) % 2 == 1
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            if use_alt:
                cell.fill = alt_fill
            cell.border = border
            cell.alignment = Alignment(vertical="center")


def _auto_fit_columns(ws) -> None:
    """Ajusta el ancho de columnas automáticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_sheet(
    writer: pd.ExcelWriter,
    df: pd.DataFrame,
    sheet_name: str,
    title: str,
) -> None:
    """Escribe un DataFrame en una hoja de Excel con formato completo."""
    # Fila 1: título, fila 2: vacía, fila 3: cabecera, fila 4+: datos
    df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)

    wb = writer.book
    ws = writer.sheets[sheet_name]

    # ── Título ───────────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}2")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(color=COLOR_HEADER_FG, bold=True, size=13)
    title_cell.fill = PatternFill("solid", fgColor=COLOR_TITLE_BG)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 22

    # ── Estilo de cabecera y datos ────────────────────────────────────────────
    _apply_header_style(ws, row=4, num_cols=len(df.columns))
    _apply_data_rows_style(
        ws,
        start_row=5,
        end_row=4 + len(df),
        num_cols=len(df.columns),
    )
    _auto_fit_columns(ws)

    # Congelar cabecera
    ws.freeze_panes = "A5"


def generate_excel_report(csv_path: str, output_path: str) -> None:
    """
    Lee el CSV de inventario y genera un informe Excel con varias hojas:
    - Resumen ejecutivo
    - Servidores vulnerables
    - Servidores Windows
    - Servidores con RAM baja
    - Agrupación por departamento

    Args:
        csv_path: Ruta al CSV de inventario.
        output_path: Ruta de salida del archivo Excel.
    """
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    df = load_inventory(csv_path)
    df_vulnerable = filter_vulnerable_servers(df)
    df_windows = filter_windows_servers(df)
    df_low_ram = filter_low_ram(df, threshold_gb=4)
    df_dept = group_by_department(df)

    # ── Hoja de resumen ───────────────────────────────────────────────────────
    resumen_data = {
        "Métrica": [
            "Total de servidores",
            "Servidores Windows Server",
            "Servidores Linux",
            "Servidores con < 4 GB RAM",
            "Servidores vulnerables / desactualizados",
            "Servidores en estado 'degradado'",
            "RAM media (GB)",
            "Fecha del informe",
        ],
        "Valor": [
            len(df),
            len(df_windows),
            len(df) - len(df_windows),
            len(df_low_ram),
            len(df_vulnerable),
            int((df["status"] == "degradado").sum()),
            round(df["ram_gb"].mean(), 1),
            datetime.now().strftime("%Y-%m-%d %H:%M"),
        ],
    }
    df_resumen = pd.DataFrame(resumen_data)

    print(f"\n  📊 Generando informe Excel en: {output_path}")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_sheet(writer, df_resumen, "Resumen Ejecutivo", "📊 Resumen del Inventario de Red")
        _write_sheet(writer, df_vulnerable, "Servidores Vulnerables", "🚨 Servidores Vulnerables o Desactualizados")
        _write_sheet(writer, df_windows, "Servidores Windows", "🪟 Servidores Windows Server")
        _write_sheet(writer, df_low_ram, "RAM Baja", "💾 Servidores con RAM < 4 GB")
        _write_sheet(writer, df_dept, "Por Departamento", "🏢 Distribución por Departamento")

    print(f"  ✅ Informe generado: {output_path}")
    print(f"  📋 Hojas incluidas: Resumen Ejecutivo, Servidores Vulnerables, Windows, RAM Baja, Por Departamento")
